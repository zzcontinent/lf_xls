#!python
# -*- coding: UTF-8 -*-

import datetime
import random
import sys
import os
import openpyxl
import math
import statistics
import numpy as np

xls_template = './res/template.xlsx'

#===================================================通用
#pre0 生成随机数
def gen_random_range(left,right,round_cnt=None):
    if round_cnt:
        return round(random.uniform(left, right),round_cnt)
    else:
        return random.uniform(left, right)

#pre1 公式生存随机数(偏差小)

#pre2 把合格随机数写入xlsx
def open_xls(file_name):
    global wb
    global ws
    wb = openpyxl.load_workbook(xls_template, data_only=False)
    ws = wb[wb.sheetnames[0]]

def write_xls(row,col,data):
    ws.cell(row, col, data)
    wb.save(xls_template)

def read_xls(row, col):
    return ws.cell(row, col).value

#pre3 公式
def test_formula():
    print(round(1.23123,3))
    print(math.log(10,2))
    print(statistics.stdev([1,2,3]))
    print(np.mean([1,2,3]))
    print(np.std([1,2,3]))
    print(np.var([1,2,3]))

# ===================================================具体项
#0.主曲线+双对数

#1. 定标
def calc_out1(in_b16, in_b20, in_b21, in_b22, in_c16, in_c20, in_c21, in_c22, in_e6, in_e7, in_e8, in_e9, in_e10):
    mid_b23=(in_b20 + in_b21 + in_b22)/3
    mid_c23=(in_c20 + in_c21 + in_c22)/3
    mid_b17=round((10**(((in_e7 - in_e6)/(((math.log10(in_b16*1000)*1000/in_e8)**(-in_e9)+1)**in_e10)+in_e6)/1000)), 0)
    mid_c17=round((10**(((in_e7 - in_e6)/(((math.log10(in_c16*1000)*1000/in_e8)**(-in_e9)+1)**in_e10)+in_e6)/1000)), 0)
    out_b25=round((mid_b23 - mid_b17)/ mid_b17, 4)
    out_c25=round((mid_c23 - mid_c17)/ mid_c17, 4)
    print("=============")
    print(mid_b17,mid_c17)
    print("=============")
    print(mid_b23,mid_c23)
    print("=============")
    print(out_b25,out_c25)
    print("=============")
    return out_b25, out_c25

#2. 准确度
def calc_out2(in_a34, in_c34, in_c35, in_c36):
    mid_c37=np.mean([in_c34, in_c35, in_c36])
    mid_c38=np.std([in_c34, in_c35, in_c36])
    out_c39=mid_c38/mid_c37
    out_c40=(mid_c37 - in_a34)/in_a34
    print("=============")
    print(out_c39, out_c40)
    mid_d34=round((10**(((in_e7 - in_e6)/(((math.log10(in_c34*1000)*1000/in_e8)**(-in_e9)+1)**in_e10)+in_e6)/1000)), 0)
    mid_d35=round((10**(((in_e7 - in_e6)/(((math.log10(in_c35*1000)*1000/in_e8)**(-in_e9)+1)**in_e10)+in_e6)/1000)), 0)
    mid_d36=round((10**(((in_e7 - in_e6)/(((math.log10(in_c36*1000)*1000/in_e8)**(-in_e9)+1)**in_e10)+in_e6)/1000)), 0)
    print("=============")
    print(mid_d34, mid_d35, mid_d36)
    mid_d37=np.mean([mid_d34, mid_d35, mid_d36])
    mid_d38=np.std([mid_d34, mid_d35, mid_d36])
    mid_d39=mid_d38/mid_d37
    print("=============")
    print(mid_d37, mid_d38, mid_d39)
    


#3. 最低检测限


#4. 线性


#5. 批内及批间精密度


#6. 特异性

#7. 批间差

#8. 校准品
#8.1 校准品均一性
#8.1.1 同批号五瓶校

#8.1.2 同一瓶五次测定

#8.2 校准品准确性


def auto_run():
    open_xls(xls_template)
    print(math.log(10,2))
    print(read_xls(2,1))

# test_formula()
# auto_run()
# def calc_out1(in_b16, in_b20, in_b21, in_b22, in_c16, in_c20, in_c21, in_c22, in_e6, in_e7, in_e8, in_e9, in_e10):
open_xls(xls_template)
calc_out1(read_xls(16,2), read_xls(20,2),read_xls(21,2), read_xls(22,2),
        read_xls(16,3), read_xls(20,3),read_xls(21,3),read_xls(22,3),
        read_xls(6,5),read_xls(7,5),read_xls(8,5),read_xls(9,5),read_xls(10,5))
# calc_out1(5.04, 40226, 39402,39328,    101.7, 645892,636546,625769,    1444.3381,64304.5915,61377.8843,4.75075,0.22387)
#auto_run()

